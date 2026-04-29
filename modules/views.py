"""
Views for the modules app.
Endpoints are nested under /api/v1/projects/<project_id>/modules/.
"""

import io
import logging
from datetime import datetime

from django.db import transaction
from rest_framework import status
from rest_framework.exceptions import NotFound, PermissionDenied
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from notifications.models import NotificationType
from notifications.services import notify_project_related_users
from rbac.permissions import IsActiveUser
from projects.models import Project

from .models import Module
from .serializers import ModuleSerializer
from .permissions import can_manage_module

logger = logging.getLogger(__name__)


def _get_project(pk):
    try:
        return Project.objects.select_related('owner').get(pk=pk)
    except Project.DoesNotExist:
        raise NotFound('Project not found.')


def _check_project_visibility(user, project):
    # Project visibility is org-wide for active authenticated users.
    return True


class ModuleListCreateView(APIView):
    """
    GET  /api/v1/projects/<project_id>/modules/
    POST /api/v1/projects/<project_id>/modules/
    """
    permission_classes = [IsAuthenticated, IsActiveUser]

    def get(self, request, project_id):
        project = _get_project(project_id)
        _check_project_visibility(request.user, project)

        qs = Module.objects.filter(project=project).prefetch_related('tasks')

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        paginator = PageNumberPagination()
        paginator.page_size = 50
        page = paginator.paginate_queryset(qs, request)
        serializer = ModuleSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request, project_id):
        project = _get_project(project_id)
        if not can_manage_module(request.user, project):
            raise PermissionDenied('Only the project owner or admin can create modules.')

        data = dict(request.data)
        data['project'] = str(project.pk)

        serializer = ModuleSerializer(data=data, context={'request': request, 'project': project})
        if not serializer.is_valid():
            return Response(
                {'detail': 'Validation failed.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        module = serializer.save(created_by=request.user, updated_by=request.user)
        try:
            notify_project_related_users(
                project=project,
                notification_type=NotificationType.GENERAL,
                title=f'Module created: {module.name}',
                body=f"Module '{module.name}' was created in project '{project.name}'.",
                link=f'/projects/{project.pk}/modules/{module.pk}/',
                exclude_user_ids=[request.user.id],
            )
        except Exception:
            pass
        logger.info('%s created module %s in project %s', request.user.email, module.name, project.name)
        return Response(ModuleSerializer(module).data, status=status.HTTP_201_CREATED)


class ModuleDetailView(APIView):
    """
    GET    /api/v1/projects/<project_id>/modules/<id>/
    PUT    /api/v1/projects/<project_id>/modules/<id>/
    PATCH  /api/v1/projects/<project_id>/modules/<id>/
    DELETE /api/v1/projects/<project_id>/modules/<id>/
    """
    permission_classes = [IsAuthenticated, IsActiveUser]

    def _get_module(self, project_id, pk):
        project = _get_project(project_id)
        try:
            module = Module.objects.get(pk=pk, project=project)
        except Module.DoesNotExist:
            raise NotFound('Module not found.')
        return project, module

    def get(self, request, project_id, pk):
        project, module = self._get_module(project_id, pk)
        _check_project_visibility(request.user, project)
        return Response(ModuleSerializer(module).data)

    def _update(self, request, project_id, pk, partial):
        project, module = self._get_module(project_id, pk)
        if not can_manage_module(request.user, project):
            raise PermissionDenied('Only the project owner or admin can edit modules.')
        previous_status = module.status

        data = dict(request.data)
        data['project'] = str(project.pk)

        serializer = ModuleSerializer(module, data=data, partial=partial, context={'request': request, 'project': project})
        if not serializer.is_valid():
            return Response(
                {'detail': 'Validation failed.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        module = serializer.save(updated_by=request.user)
        try:
            if previous_status != module.status:
                notify_project_related_users(
                    project=project,
                    notification_type=NotificationType.STATUS_CHANGED,
                    title=f'Module status changed: {module.name}',
                    body=(
                        f"Module '{module.name}' status changed from "
                        f"'{previous_status}' to '{module.status}'."
                    ),
                    link=f'/projects/{project.pk}/modules/{module.pk}/',
                    exclude_user_ids=[request.user.id],
                )
            else:
                notify_project_related_users(
                    project=project,
                    notification_type=NotificationType.GENERAL,
                    title=f'Module updated: {module.name}',
                    body=f"Module '{module.name}' was updated in project '{project.name}'.",
                    link=f'/projects/{project.pk}/modules/{module.pk}/',
                    exclude_user_ids=[request.user.id],
                )
        except Exception:
            pass
        logger.info('%s updated module %s', request.user.email, module.name)
        return Response(ModuleSerializer(module).data)

    def put(self, request, project_id, pk):
        return self._update(request, project_id, pk, partial=False)

    def patch(self, request, project_id, pk):
        return self._update(request, project_id, pk, partial=True)

    def delete(self, request, project_id, pk):
        project, module = self._get_module(project_id, pk)
        if not can_manage_module(request.user, project):
            raise PermissionDenied('Only the project owner or admin can delete modules.')
        name = module.name
        module.delete()
        try:
            notify_project_related_users(
                project=project,
                notification_type=NotificationType.GENERAL,
                title=f'Module deleted: {name}',
                body=f"Module '{name}' was deleted from project '{project.name}'.",
                link=f'/projects/{project.pk}/modules/',
                exclude_user_ids=[request.user.id],
            )
        except Exception:
            pass
        logger.info('%s deleted module %s', request.user.email, name)
        return Response({'detail': f"Module '{name}' deleted."})


# ---------------------------------------------------------------------------
# Excel bulk import
# ---------------------------------------------------------------------------

class ModuleBulkImportView(APIView):
    """
    POST /api/v1/projects/<project_id>/modules/bulk-import/
    GET  /api/v1/projects/<project_id>/modules/bulk-import/template/
    """
    permission_classes = [IsAuthenticated, IsActiveUser]

    def get(self, request, project_id):
        """Return a sample .xlsx template."""
        try:
            import openpyxl
        except ImportError:
            return Response({'detail': 'openpyxl is not installed.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        from django.http import HttpResponse

        wb = openpyxl.Workbook()

        # Modules sheet
        ws_modules = wb.active
        ws_modules.title = 'Modules'
        ws_modules.append(['module_name', 'module_description', 'module_status', 'module_deadline'])
        ws_modules.append(['Example Module', 'Module description here', 'todo', '2025-12-31'])

        # Tasks sheet
        ws_tasks = wb.create_sheet('Tasks')
        ws_tasks.append(['module_name', 'task_title', 'task_description', 'task_status', 'task_priority', 'task_due_date'])
        ws_tasks.append(['Example Module', 'Example Task', 'Task description', 'todo', 'medium', '2025-12-15'])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="modules_tasks_template.xlsx"'
        return response

    def post(self, request, project_id):
        """Validate and import modules + tasks from .xlsx file."""
        try:
            import openpyxl
        except ImportError:
            return Response({'detail': 'openpyxl is not installed.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        project = _get_project(project_id)
        if not can_manage_module(request.user, project):
            raise PermissionDenied('Only the project owner or admin can bulk import modules.')

        file = request.FILES.get('file')
        if not file:
            return Response(
                {'detail': 'No file provided.', 'errors': [{'row': None, 'sheet': None, 'errors': ['file field is required.']}]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not file.name.endswith('.xlsx'):
            return Response(
                {'detail': 'Invalid file type.', 'errors': [{'row': None, 'sheet': None, 'errors': ['Only .xlsx files are accepted.']}]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            return Response(
                {'detail': 'Could not read Excel file.', 'errors': [{'row': None, 'sheet': None, 'errors': [str(e)]}]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        errors = []
        modules_data = []
        tasks_data = []

        # --- Parse Modules sheet ---
        if 'Modules' in wb.sheetnames:
            ws = wb['Modules']
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                pass  # No data rows — OK, empty modules sheet
            else:
                header = [str(h).strip().lower() if h else '' for h in rows[0]]
                required_module_cols = {'module_name'}
                for col in required_module_cols:
                    if col not in header:
                        errors.append({'row': 1, 'sheet': 'Modules', 'errors': [f"Missing required column: '{col}'"]})

                if not errors:
                    for i, row in enumerate(rows[1:], start=2):
                        row_data = dict(zip(header, row))
                        row_errors = []

                        name = str(row_data.get('module_name') or '').strip()
                        if not name:
                            row_errors.append('module_name is required.')

                        mod_status = str(row_data.get('module_status') or 'todo').strip().lower()
                        valid_statuses = ['todo', 'in_progress', 'blocked', 'done']
                        if mod_status not in valid_statuses:
                            row_errors.append(f"module_status must be one of: {', '.join(valid_statuses)}")

                        deadline = None
                        raw_deadline = row_data.get('module_deadline')
                        if raw_deadline:
                            if hasattr(raw_deadline, 'date'):
                                deadline = raw_deadline.date()
                            else:
                                try:
                                    deadline = datetime.strptime(str(raw_deadline).strip(), '%Y-%m-%d').date()
                                except ValueError:
                                    row_errors.append("module_deadline must be YYYY-MM-DD format.")

                        if row_errors:
                            errors.extend([{'row': i, 'sheet': 'Modules', 'errors': row_errors}])
                        else:
                            modules_data.append({
                                'name': name,
                                'description': str(row_data.get('module_description') or '').strip(),
                                'status': mod_status,
                                'deadline': deadline,
                            })

        # --- Parse Tasks sheet ---
        if 'Tasks' in wb.sheetnames:
            ws = wb['Tasks']
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) >= 2:
                header = [str(h).strip().lower() if h else '' for h in rows[0]]
                required_task_cols = {'module_name', 'task_title'}
                for col in required_task_cols:
                    if col not in header:
                        errors.append({'row': 1, 'sheet': 'Tasks', 'errors': [f"Missing required column: '{col}'"]})

                if not [e for e in errors if e.get('sheet') == 'Tasks' and e.get('row') == 1]:
                    for i, row in enumerate(rows[1:], start=2):
                        row_data = dict(zip(header, row))
                        row_errors = []

                        mod_name = str(row_data.get('module_name') or '').strip()
                        if not mod_name:
                            row_errors.append('module_name is required.')

                        task_title = str(row_data.get('task_title') or '').strip()
                        if not task_title:
                            row_errors.append('task_title is required.')

                        task_status = str(row_data.get('task_status') or 'todo').strip().lower()
                        valid_statuses = ['todo', 'in_progress', 'blocked', 'done']
                        if task_status not in valid_statuses:
                            row_errors.append(f"task_status must be one of: {', '.join(valid_statuses)}")

                        task_priority = str(row_data.get('task_priority') or 'medium').strip().lower()
                        valid_priorities = ['low', 'medium', 'high', 'critical']
                        if task_priority not in valid_priorities:
                            row_errors.append(f"task_priority must be one of: {', '.join(valid_priorities)}")

                        due_date = None
                        raw_due = row_data.get('task_due_date')
                        if raw_due:
                            if hasattr(raw_due, 'date'):
                                due_date = raw_due.date()
                            else:
                                try:
                                    due_date = datetime.strptime(str(raw_due).strip(), '%Y-%m-%d').date()
                                except ValueError:
                                    row_errors.append("task_due_date must be YYYY-MM-DD format.")

                        if row_errors:
                            errors.extend([{'row': i, 'sheet': 'Tasks', 'errors': row_errors}])
                        else:
                            tasks_data.append({
                                'module_name': mod_name,
                                'title': task_title,
                                'description': str(row_data.get('task_description') or '').strip(),
                                # Assignee is intentionally left unassigned in Excel imports.
                                # Project contributors can assign from the frontend after import.
                                'assignee': None,
                                'status': task_status,
                                'priority': task_priority,
                                'due_date': due_date,
                            })

        if errors:
            return Response({'detail': 'Import failed due to validation errors.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        if not modules_data and not tasks_data:
            return Response({'detail': 'No data found in file.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate that task module_names match modules being imported or existing modules
        existing_module_names = set(Module.objects.filter(project=project).values_list('name', flat=True))
        importing_module_names = {m['name'] for m in modules_data}
        all_known_names = existing_module_names | importing_module_names

        for i, task in enumerate(tasks_data):
            if task['module_name'] not in all_known_names:
                errors.append({
                    'row': i + 2,
                    'sheet': 'Tasks',
                    'errors': [f"module_name '{task['module_name']}' not found in this project."],
                })

        if errors:
            return Response({'detail': 'Import failed due to validation errors.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Save in a transaction
        from tasks.models import Task

        modules_created = 0
        tasks_created = 0
        module_map = {}

        with transaction.atomic():
            # Create/update modules
            for mod_data in modules_data:
                module, created = Module.objects.get_or_create(
                    project=project,
                    name=mod_data['name'],
                    defaults={
                        'description': mod_data['description'],
                        'status': mod_data['status'],
                        'deadline': mod_data['deadline'],
                        'created_by': request.user,
                        'updated_by': request.user,
                    },
                )
                if created:
                    modules_created += 1
                module_map[mod_data['name']] = module

            # Populate map with existing modules not being re-created
            for name in existing_module_names:
                if name not in module_map:
                    try:
                        module_map[name] = Module.objects.get(project=project, name=name)
                    except Module.DoesNotExist:
                        pass

            # Create tasks
            for task_data in tasks_data:
                module = module_map.get(task_data['module_name'])
                if module:
                    Task.objects.create(
                        module=module,
                        title=task_data['title'],
                        description=task_data['description'],
                        assignee=task_data['assignee'],
                        status=task_data['status'],
                        priority=task_data['priority'],
                        due_date=task_data['due_date'],
                        created_by=request.user,
                        updated_by=request.user,
                    )
                    tasks_created += 1

        logger.info('%s bulk-imported %d modules and %d tasks into project %s',
                    request.user.email, modules_created, tasks_created, project.name)

        return Response(
            {'modules_created': modules_created, 'tasks_created': tasks_created},
            status=status.HTTP_201_CREATED,
        )
